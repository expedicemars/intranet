from openpyxl import Workbook
from website.paths import exporty_path, mailing_list_path, pohovory_path, poznamky_path, velitel_odbornosti_data_path, zadani_folder_path, admin_logs_file_path, app_logs_file_path, prohlaseni_path, user_data_folder_path, prubeh_rocniku_path, sablony_folder_path, odkazy_path
from website.helpers.pretty_date import pretty_date, pretty_datetime
from website.json_handlers.dostupne_omezeni import get_dostupne_odbornosti
from website.role_handler import get_access_rights
from datetime import datetime, date
from website.models.user import User
from website.models.hodnoceni import Hodnoceni
from website import db
import json
from zipfile import ZipFile

def exportovat() -> None:
    time_of_export = datetime.now()
    filename = time_of_export.strftime("%Y_%m_%d-%H_%M_%S")
    folder_name = time_of_export.isoformat()
    folder = exporty_path() / folder_name
    folder.mkdir()
    xlsx_path = folder / "data.xlsx"
    zip_filename = "zaloha_iEM_" + filename + ".zip"
    zip_path = folder / zip_filename

    wb = Workbook()
    wb.remove(wb.active)
    ws1 = wb.create_sheet("Účastníci")
    names = [
        "ID",
        "E-mail",
        "confirmed",
        "Jméno",
        "Adresa",
        "Telefonní číslo",
        "E-mail rodičů",
        "Odbornost",
        "Datum narození",
        "Progress",
        "Role",
        "Tričko",
        "Jak se o nás dozvěděli",
        "Admin poznámka",
        "Uzamčené změny",
        "Alergie",
        "Škola",
        "Datum registrace",
        "Datum pohovoru",
        "Meeting link",
        "Odevzdal motivační dotazník",
        "Oslovení v 1. pádě",
        "Oslovení v 5. pádě",
        "Zájmeno"
        ]
    for i, name in enumerate(names):
        ws1.cell(1,i+1, value=name)
    for i, u in enumerate(User.get_all()):
        ws1.cell(i+2,1,value=u.id)
        ws1.cell(i+2,2,value=u.email)
        ws1.cell(i+2,3,value=u.confirmed)
        ws1.cell(i+2,4,value=u.jmeno)
        ws1.cell(i+2,5,value=u.adresa)
        ws1.cell(i+2,6,value=u.telcislo)
        ws1.cell(i+2,7,value=u.mail_rodicu)
        ws1.cell(i+2,8,value=u.odbornost)
        ws1.cell(i+2,9,value=pretty_date(u.datum_narozeni))
        ws1.cell(i+2,10,value=u.progress)
        ws1.cell(i+2,11,value=u.role)
        ws1.cell(i+2,12,value=u.tricko)
        ws1.cell(i+2,13,value=u.dozvedeli)
        ws1.cell(i+2,14,value=u.admin_poznamka)
        ws1.cell(i+2,15,value=u.uzamcene_zmeny)
        ws1.cell(i+2,16,value=u.alergie)
        ws1.cell(i+2,17,value=u.skola)
        ws1.cell(i+2,18,value=pretty_datetime(u.datum_registrace))
        ws1.cell(i+2,19,value=pretty_datetime(u.datum_pohovoru))
        ws1.cell(i+2,20,value=u.meeting_link)
        ws1.cell(i+2,21,value=u.odevzdany_motivacni_dotaznik)
        ws1.cell(i+2,22,value=u.osloveni_1p)
        ws1.cell(i+2,23,value=u.osloveni_5p)
        ws1.cell(i+2,24,value=u.zajmeno)

    ws2 = wb.create_sheet("Pohovory")
    with open(pohovory_path()) as file:
        file = json.load(file)
    ws2.cell(1,1,"Datum")
    ws2.cell(1,2,"ID účastníka")
    ws2.cell(1,3,"Kdo to vypsal")
    for i, zaznam in enumerate(file):
        ws2.cell(i+2,1,pretty_datetime(zaznam["iso"]))
        ws2.cell(i+2,2,zaznam["user"])
        ws2.cell(i+2,3,zaznam["admin"])
    
    ws3 = wb.create_sheet("Poznámky")
    with open(poznamky_path()) as file:
        file = json.load(file)
    ws3.cell(1,1,"Autor")
    ws3.cell(1,2,"Datum")
    ws3.cell(1,3,"Poznámka")
    for i, zaznam in enumerate(file):
        ws3.cell(i+2,1,zaznam["autor"])
        ws3.cell(i+2,2,zaznam["datum"])
        ws3.cell(i+2,3,zaznam["msg"])
    
    
    ws4 = wb.create_sheet("Kontakty na velitele odborností")
    with open(velitel_odbornosti_data_path()) as file:
        file = json.load(file)
    ws4.cell(1,1,"Odbornost")
    ws4.cell(1,2,"Kontakt")
    for i, key in enumerate(file.keys()):
        ws4.cell(i+2,1,key)
        ws4.cell(i+2,2,file[key])
        
    ws5 = wb.create_sheet("Hodnocení")
    ws5.cell(1,1,"ID")
    ws5.cell(1,2,"ID uživatele")
    ws5.cell(1,3,"Věcnost")
    ws5.cell(1,4,"Originalita")
    ws5.cell(1,5,"Komunikace")
    ws5.cell(1,6,"Motivovanost")
    ws5.cell(1,7,"Sebevědomí")
    ws5.cell(1,8,"Flexibilita")
    ws5.cell(1,9,"Sebehodnocení")
    ws5.cell(1,10,"K-faktor")
    ws5.cell(1,11,"Dojem")
    ws5.cell(1,12,"Datum založení")
    ws5.cell(1,13,"ID admina")
    for i, hodnoceni in enumerate(Hodnoceni.get_all(), start=2):
        ws5.cell(i, 1, hodnoceni.id)
        ws5.cell(i, 2, hodnoceni.user_id)
        ws5.cell(i, 3, hodnoceni.vecnost)
        ws5.cell(i, 4, hodnoceni.originalita)
        ws5.cell(i, 5, hodnoceni.komunikace)
        ws5.cell(i, 6, hodnoceni.motivovanost)
        ws5.cell(i, 7, hodnoceni.sebevedomi)
        ws5.cell(i, 8, hodnoceni.flexibilita)
        ws5.cell(i, 9, hodnoceni.sebehodnoceni)
        ws5.cell(i, 10, hodnoceni.k_faktor)
        ws5.cell(i, 11, hodnoceni.dojem)
        ws5.cell(i, 12, hodnoceni.datum_zalozeni)
        ws5.cell(i, 13, hodnoceni.admin_id)

    wb.save(xlsx_path)

    with ZipFile(zip_path, mode="a") as archive:
        archive.write(xlsx_path, arcname=xlsx_path.name)
        archive.write(admin_logs_file_path(), arcname=admin_logs_file_path().name)
        archive.write(app_logs_file_path(), arcname=app_logs_file_path().name)
        archive.write(prohlaseni_path(), arcname=prohlaseni_path().name)
        archive.write(odkazy_path(), arcname=odkazy_path().name)
        for file in sablony_folder_path().iterdir():
            archive.write(file, arcname=file.name)
        archive.write(prubeh_rocniku_path(), arcname=prubeh_rocniku_path().name)
        archive.write(mailing_list_path(), arcname=mailing_list_path().name)
        for p in user_data_folder_path().rglob("*"):
            archive.write(p, arcname=p.relative_to(user_data_folder_path().parent))
        for p in zadani_folder_path().rglob("*"):
            archive.write(p, arcname=p.relative_to(zadani_folder_path().parent))

def promazat() -> None:
    """
    Promaže vše, co bylo exportováno, krom mailing listu, prohlášení pro rodiče, šablon a app logů.
    """
    for folder in zadani_folder_path().iterdir():
        if folder.name == ".DS_Store":
            pass
        else:
            for file in folder.iterdir():
                if file.name == ".DS_Store":
                    pass
                else:
                    file.unlink()
    with open(admin_logs_file_path(), "w") as file:
        file.write("")
    with open(pohovory_path(), "w") as file:
        file.write(json.dumps([], indent=4))
    with open(poznamky_path(), "w") as file:
        file.write(json.dumps([], indent=4))
    with open(prubeh_rocniku_path(),"w") as file:
        file.write(json.dumps({"datum_konce_registrace":str(date.today()),"datum_zacatku_registrace":str(date.today()),"otevrena_registrace":False,"viditelna_zadani":False, "koordinator_internetovych_kol":""}, indent=4))
    with open(velitel_odbornosti_data_path(), "w") as file:
        file.write(json.dumps({odb["system_name"]:"" for odb in get_dostupne_odbornosti()}, indent=4))
    
    for h in Hodnoceni.get_all():
        db.session.delete(h)
        db.session.commit()
        
    for u in User.get_all():
        u: User
        if "admin" not in get_access_rights(u):
            u.odstranit()
        else:
            u.datum_pohovoru = None
            db.session.add(u)
            db.session.commit()
    
        


    
